#!/usr/bin/env python3
"""
Build the RTB Apprenticeship Useful Information site from the Excel workbook.

Usage:  python build.py Useful_Information_for_Apprenticeships.xlsx
Output: index.html  (a single self-contained file ready for GitHub Pages)

Workbook layout this script expects (per sheet):
  Row 1 : tab title in col A; col C may carry the tab-level SharePoint folder link
  Row 2 : blank
  Row 3 : header row  ->  #  |  Subfolder  |  Notes / Link
  Row 4+: data rows   ->  number | subfolder name | notes/link cell

The SharePoint URLs are stored as *embedded hyperlinks* on the col C cells,
not as visible text. This script reads the hyperlink target where present and
falls back to a "pending" flag where it is not.

To update the live site: add/adjust hyperlinks in the spreadsheet, re-run this
script, and commit the regenerated index.html.
"""

import sys, json, datetime
import openpyxl

SKIP_SHEETS = set()
DATA_START = 4  # first data row (row 3 is the header)


def clean(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%d %b %Y")
    return str(v).strip()


def load(path):
    wb = openpyxl.load_workbook(path)  # keep hyperlinks (no data_only)
    sheets = []
    for ws in wb.worksheets:
        if ws.title in SKIP_SHEETS:
            continue
        title_cell = ws.cell(row=1, column=3)
        folder_link = title_cell.hyperlink.target if title_cell.hyperlink else ""

        rows = []
        for r in range(DATA_START, ws.max_row + 1):
            subfolder = clean(ws.cell(row=r, column=2).value)
            link_cell = ws.cell(row=r, column=3)
            link = link_cell.hyperlink.target if link_cell.hyperlink else ""
            note = clean(link_cell.value)
            if subfolder == "" and note == "" and link == "":
                continue
            rows.append({"name": subfolder, "note": note, "link": link})

        sheets.append({"name": ws.title.strip(), "folder": folder_link, "rows": rows})
    return sheets


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else "Useful_Information_for_Apprenticeships.xlsx"
    sheets = load(src)
    built = datetime.datetime.now().strftime("%d %B %Y, %H:%M")
    data_json = json.dumps({"sheets": sheets, "built": built})
    htmlout = TEMPLATE.replace("/*DATA*/", data_json)
    with open("index.html", "w", encoding="utf-8") as f:
        f.write(htmlout)
    total = sum(len(s["rows"]) for s in sheets)
    linked = sum(1 for s in sheets for r in s["rows"] if r["link"])
    print(f"Built index.html: {len(sheets)} tabs, {total} rows, {linked} live links.")


TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>RTB Apprenticeship - Useful Information</title>
<style>
  :root{
    --navy:#1b2a4a; --navy2:#25396b; --accent:#e8552d;
    --ink:#1f2733; --muted:#63708a; --line:#e3e8f0; --bg:#f4f6fa; --card:#fff;
    --pending:#c0392b; --pendbg:#fdecea;
  }
  *{box-sizing:border-box}
  body{margin:0;font-family:Arial,Helvetica,sans-serif;color:var(--ink);background:var(--bg);line-height:1.45;font-size:16px}
  header{background:linear-gradient(120deg,var(--navy),var(--navy2));color:#fff;padding:22px 26px}
  header h1{margin:0;font-size:23px}
  header p{margin:6px 0 0;color:#c5d0e6;font-size:15px}
  .wrap{max-width:1150px;margin:0 auto;padding:0 18px}
  .toolbar{position:sticky;top:0;z-index:5;background:var(--bg);padding:14px 0 8px;border-bottom:1px solid var(--line)}
  .search{width:100%;padding:12px 14px;font-size:16px;border:1px solid var(--line);border-radius:8px}
  .tabs{display:flex;flex-wrap:wrap;gap:6px;margin-top:12px}
  .tab{border:1px solid var(--line);background:var(--card);color:var(--navy);padding:8px 13px;border-radius:20px;font-size:15px;cursor:pointer;white-space:nowrap}
  .tab .n{color:var(--muted);font-size:15px;margin-left:5px}
  .tab.active{background:var(--navy);color:#fff;border-color:var(--navy)}
  .tab.active .n{color:#aebbd6}
  .tab.audit{border-color:var(--accent);color:var(--accent)}
  .tab.audit.active{background:var(--accent);color:#fff;border-color:var(--accent)}
  main{padding:20px 0 60px}
  .sheet-head{display:flex;align-items:center;gap:14px;flex-wrap:wrap;margin:6px 0 14px}
  .sheet-title{font-size:19px;margin:0;color:var(--navy)}
  .folder-link{background:var(--accent);color:#fff;text-decoration:none;padding:7px 13px;border-radius:6px;font-size:15px;font-weight:bold}
  .folder-link:hover{background:#cf4522}
  table{width:100%;border-collapse:collapse;background:var(--card);border-radius:10px;overflow:hidden;box-shadow:0 1px 3px rgba(20,30,60,.06)}
  th{background:var(--navy);color:#fff;text-align:left;padding:12px 14px;font-size:15px;text-transform:uppercase;letter-spacing:.4px;font-weight:bold}
  td{padding:12px 14px;border-top:1px solid var(--line);font-size:16px;vertical-align:top}
  tr:hover td{background:#fafbfe}
  td.linkcell a{color:#fff;background:var(--navy2);text-decoration:none;padding:6px 12px;border-radius:6px;font-size:15px;font-weight:bold;display:inline-block}
  td.linkcell a:hover{background:var(--navy)}
  .note{color:var(--muted);font-size:14px;margin-left:8px}
  .pill{display:inline-block;padding:4px 10px;border-radius:20px;font-size:14px;font-weight:bold;white-space:nowrap}
  .pill.pending{background:var(--pendbg);color:var(--pending)}
  .empty{padding:26px;text-align:center;color:var(--muted);background:var(--card);border-radius:10px;font-size:16px}
  .count{color:var(--muted);font-size:15px;margin-bottom:10px}
  footer{color:var(--muted);font-size:14px;text-align:center;padding:22px}
  @media(max-width:640px){th,td{padding:9px;font-size:15px}header h1{font-size:20px}}
</style>
</head>
<body>
<header>
  <div class="wrap">
    <h1>Apprenticeship &mdash; Useful Information</h1>
    <p>Central directory of resource folders and links for the apprenticeship team.</p>
  </div>
</header>

<div class="wrap">
  <div class="toolbar">
    <input class="search" id="search" type="text" placeholder="Search across all tabs...">
    <div class="tabs" id="tabs"></div>
  </div>
  <main id="main"></main>
</div>

<footer id="footer"></footer>

<script>
const DATA = /*DATA*/;
function esc(s){ return (s||'').replace(/[&<>"]/g, c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c])); }

let activeTab = 0;
let query = '';

function rowMatches(r){
  if(!query) return true;
  return (r.name+' '+r.note).toLowerCase().includes(query);
}

function linkCell(r){
  if(r.link){
    return '<a href="'+esc(r.link)+'" target="_blank" rel="noopener">Open folder &#8599;</a>'
      + (r.note && r.note.toLowerCase()!==r.name.toLowerCase() ? '<span class="note">'+esc(r.note)+'</span>' : '');
  }
  return '<span class="pill pending">Pending link</span>'
    + (r.note ? '<span class="note">'+esc(r.note)+'</span>' : '');
}

function renderSheet(s){
  const rows = s.rows.filter(rowMatches);
  let h = '<div class="sheet-head"><h2 class="sheet-title">'+esc(s.name)+'</h2>';
  if(s.folder) h += '<a class="folder-link" href="'+esc(s.folder)+'" target="_blank" rel="noopener">Open whole folder &#8599;</a>';
  h += '</div>';
  h += '<div class="count">'+rows.length+' of '+s.rows.length+' item'+(s.rows.length===1?'':'s')+'</div>';
  if(rows.length===0) return h+'<div class="empty">No matching items on this tab.</div>';
  h += '<table><thead><tr><th style="width:56px">#</th><th>Subfolder</th><th style="width:38%">Link</th></tr></thead><tbody>';
  rows.forEach((r,i)=>{ h += '<tr><td>'+(i+1)+'</td><td>'+esc(r.name)+'</td><td class="linkcell">'+linkCell(r)+'</td></tr>'; });
  h += '</tbody></table>';
  return h;
}

function renderAudit(){
  const items = [];
  DATA.sheets.forEach(s=>s.rows.forEach(r=>{ if(!r.link) items.push({tab:s.name, name:r.name||'(unnamed)'}); }));
  const shown = items.filter(it=>!query || it.tab.toLowerCase().includes(query) || it.name.toLowerCase().includes(query));
  let live=0; DATA.sheets.forEach(s=>s.rows.forEach(r=>{ if(r.link) live++; }));
  let h = '<h2 class="sheet-title">Link audit</h2>';
  h += '<div class="count">'+live+' live link'+(live===1?'':'s')+' &middot; '+items.length+' still pending a URL</div>';
  if(shown.length===0) return h+'<div class="empty">Nothing pending &mdash; every subfolder has a live link. &#127881;</div>';
  h += '<table><thead><tr><th>Tab</th><th>Subfolder</th><th style="width:160px">Status</th></tr></thead><tbody>';
  shown.forEach(it=>{ h += '<tr><td>'+esc(it.tab)+'</td><td>'+esc(it.name)+'</td><td><span class="pill pending">Pending link</span></td></tr>'; });
  h += '</tbody></table>';
  return h;
}

function render(){
  document.getElementById('main').innerHTML = (activeTab===-1) ? renderAudit() : renderSheet(DATA.sheets[activeTab]);
}

function buildTabs(){
  const c = document.getElementById('tabs');
  let h = '';
  DATA.sheets.forEach((s,i)=>{ h += '<button class="tab'+(i===activeTab?' active':'')+'" data-i="'+i+'">'+esc(s.name)+'<span class="n">'+s.rows.length+'</span></button>'; });
  h += '<button class="tab audit'+(activeTab===-1?' active':'')+'" data-i="-1">Link audit</button>';
  c.innerHTML = h;
  c.querySelectorAll('.tab').forEach(b=>b.addEventListener('click',()=>{ activeTab=parseInt(b.dataset.i,10); buildTabs(); render(); }));
}

document.getElementById('search').addEventListener('input', e=>{ query=e.target.value.trim().toLowerCase(); render(); });
document.getElementById('footer').textContent = 'Raise the Bar Ltd \u00b7 Built from Useful_Information_for_Apprenticeships.xlsx on ' + DATA.built;

buildTabs();
render();
</script>
</body>
</html>
"""

if __name__ == "__main__":
    main()
